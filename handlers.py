from coroweb import get, post
from models import *
import models
import switch_tools
import config
import openpyxl
from io import BytesIO
import json

@get('/login')
def login():
    return {
        '__template__' : 'login.html'
    }

@get('/test')
def test():
    L = yield from switches.findall()
    id = []
    for each in L:
        id.append(each['id'])
    for each in id:
        record = switches_groups()
        record.switch_id = each
        record.group_id = '1'
        yield from record.save()
    return 'ok'


@get('/')
def index(**kw):

    # 获取排序参数
    order = kw.get('order', None)
    if not order is None:
        #group_name = yield from category.findall(where='category_id=%s' % order)    # 获取一级分组信息
        sub_groups = yield from left_join([category, groups]).findall(where='`category`.`id`=?', order='priority asc', args=[1])    # 获取二级分组信息

        grps = []
        # 装入已分组设备信息
        for sub_group in sub_groups:
            grp_name = sub_group['groups.name']
            switch_list = yield from left_join([quick_indicator, switches, switches_groups]).findall(where='`switches_groups`.`group_id`=?', args=[sub_group['groups.id']])
            if switch_list:
                grps.append({'grp_name':grp_name, 'switch_list':switch_list})

        # 装入未分组设备信息
        switch_list = yield from left_join([quick_indicator, switches, switches_groups]).findall(where='`switches_groups`.`group_id` is null', args=[])
        if switch_list:
            grps.append({'grp_name':'未指定分组', 'switch_list':switch_list})
    else:
        # 未指定排序则
        switch_list = yield from quick_indicator.findall()
        grps = [{'grp_name':'DEVICE INDICATORS', 'switch_list':switch_list}]

    for grp in grps:
        indicators = []
        for index, each in enumerate(grp['switch_list']):
            args = {
                'style' : 'odd_line' if (index & 0x1) else 'even_line',
                'id' : each['switch_id'],
                'host' : each['host'],
                'name' : each['name'],
                'cpus' : each['cpu'].split(),
                'memorys' : each['memory'].split(),
            }
            indicators.append(dict(**args))
        grp.pop('switch_list')
        grp.update({'indicators':indicators})

    ctgs = yield from category.findall()

    return {
        '__template__':'index.html',
        'groups' : grps,
        'categorys' : ctgs
    }

@get('/detail')
def detail(*, swid):
    switch = yield from switches.find(pk=[swid])
    extra = yield from indicator.findall(where="`switch_id` = '%s'" % swid)
    common = config.configs.common_indicator
    indicators = [key for key, value in common.items()]
    device = switch_tools.device(**switch)

    # cpu&memory
    if 'cpu' in indicators:
        cpu = device.get_graph('cpu', title='CPU Utilization')
    if 'memory' in indicators:
        memory = device.get_graph('memory', title='Memory Utilization')

    # 获取自定义指标图
    customize_indicator = []
    if not extra is None:
        for each in extra:
            name = each['name']
            argu = str(json.loads(each['argument']))
            kw = {}
            if not each['title'] is None:
                kw['title'] = each['title']
            customize_indicator.append(device.get_graph(name+'-'+argu, **kw))

    # 获取设备名字
    switch['name'] = device.get_indicator_value('name')

    return {
        '__template__' : 'detail.html',
        'cpu_img' : cpu,
        'memory_img' : memory,
        'customize_indicator' : customize_indicator,
        'switch' : switch
    }

@get('/zb.xlsx')
def zb_xlsx():
    wb = openpyxl.Workbook()

    ws = wb.active
    table = yield from quick_indicator.findall()
    # 写入首行
    for index, value in enumerate(['管理IP', '设备名称', 'CPU利用率', '内存利用率']):
        ws.cell(row=1, column=index+1).value = value
    # 写入所有数据
    seq = ['host', 'name', 'cpu', 'memory']
    for row_index, row_value in enumerate(table):
        for col_index in range(4):
            value = row_value[seq[col_index]]
            ws.cell(row=row_index+2, column=col_index+1).value = value

    f = BytesIO()
    wb.save(f)
    return f.getvalue()
    
